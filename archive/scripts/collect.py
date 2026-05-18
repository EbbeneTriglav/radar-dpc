#!/usr/bin/env python3
"""
collect.py — Raccoglitore giornaliero dati radar DPC per le aree configurate.

Per ogni area definita in archive/areas.json e per ogni prodotto (CUM24, CUM3),
scarica i GeoTIFF DPC degli ultimi N giorni, calcola statistiche dentro il
poligono dell'area + estrae il valore puntuale sui vertici campione, e
appende in modo idempotente al CSV cumulativo.

Esecuzione:
    python collect.py                  # ultimi 1 giorno
    python collect.py --days 7         # bootstrap: ultimi 7 giorni
    python collect.py --days 7 --force # bootstrap forzato (rimpiazza righe esistenti)

Output:
    archive/data/{area}_cum24.csv      append-only, idempotente per (timestamp_utc, location_name)
    archive/data/{area}_cum3.csv       idem
    archive/data/{area}.xlsx           rigenerato a fine run con 2 sheet (CUM24, CUM3)
"""

import argparse
import csv
import io
import json
import logging
import sys
import time
from datetime import datetime, timedelta, timezone
from pathlib import Path

import numpy as np
import rasterio
import rasterio.mask
import requests
from openpyxl import Workbook
from shapely.geometry import Polygon, mapping

# ─── Configurazione ───────────────────────────────────────────────────────────

DPC_API = 'https://radar-api.protezionecivile.it'
USER_AGENT = (
    'Mozilla/5.0 (Linux; X11) AppleWebKit/537.36 (KHTML, like Gecko) '
    'radar-dpc-archive/1.0'
)
HTTP_TIMEOUT = 60
HTTP_RETRIES = 3
HTTP_BACKOFF = 5

PRODUCTS = {
    'CUM24': {
        'step_minutes': 1440,        # 24h
        'utc_hours_in_day': [0],     # 1 timestamp/giorno alle 00:00 UTC
        'description': 'Cumulata 24h',
    },
    'CUM3': {
        'step_minutes': 180,         # 3h
        # Per coprire il giorno X completo serve gli 8 timestamp:
        # X 03:00 (0-3h), 06:00 (3-6), ..., 21:00 (18-21), (X+1) 00:00 (21-24)
        # Qui salviamo gli 8 timestamp etichettati come "appartenenti al giorno X"
        'utc_hours_in_day': [3, 6, 9, 12, 15, 18, 21, 24],  # 24 = giorno successivo 00:00
        'description': 'Cumulata 3h',
    },
}

CSV_HEADERS = [
    'timestamp_utc',     # ISO 8601 UTC, es. 2026-05-18T03:00:00Z
    'product',
    'area_name',
    'location_type',     # 'area' | 'vertex'
    'location_name',     # 'ruspino' / 'ruspino_v1' / ...
    'lat',               # solo per vertici
    'lon',               # solo per vertici
    'value',             # solo per vertici (mm)
    'mean',              # solo per area (mm)
    'min',               # solo per area
    'max',               # solo per area
    'pixel_count',       # solo per area
    'fetched_at_utc',
]

# ─── Setup logging ───────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%H:%M:%S',
)
log = logging.getLogger('collect')

# ─── HTTP helpers ────────────────────────────────────────────────────────────

_session = requests.Session()
_session.headers.update({'User-Agent': USER_AGENT, 'Accept': '*/*'})


def _http_request(method, url, **kwargs):
    """Richiesta HTTP con retry esponenziale."""
    kwargs.setdefault('timeout', HTTP_TIMEOUT)
    last_err = None
    for attempt in range(1, HTTP_RETRIES + 1):
        try:
            r = _session.request(method, url, **kwargs)
            if r.status_code >= 500:
                raise requests.HTTPError(f'{r.status_code} server error')
            return r
        except requests.RequestException as e:
            last_err = e
            log.warning(f'  HTTP {method} attempt {attempt}/{HTTP_RETRIES} failed: {e}')
            if attempt < HTTP_RETRIES:
                time.sleep(HTTP_BACKOFF * attempt)
    raise RuntimeError(f'HTTP failed after {HTTP_RETRIES} attempts: {last_err}')


def get_pre_signed_url(product_type, product_date_ms):
    """Chiama l'API DPC per ottenere la pre-signed URL del file S3."""
    r = _http_request('POST', f'{DPC_API}/downloadProduct',
                      json={'productType': product_type, 'productDate': product_date_ms},
                      headers={'Content-Type': 'application/json'})
    if not r.ok:
        log.warning(f'  downloadProduct {product_type} @ {product_date_ms}: HTTP {r.status_code}')
        return None
    data = r.json()
    return data.get('url')


def download_geotiff(url):
    """Scarica il GeoTIFF e ritorna i bytes (oppure None)."""
    r = _http_request('GET', url)
    if not r.ok or len(r.content) < 256:
        log.warning(f'  download fallito: HTTP {r.status_code}, {len(r.content)} bytes')
        return None
    return r.content

# ─── Calcolo statistiche su area + estrazione puntuale ───────────────────────

def stats_for_polygon(tiff_bytes, polygon_latlon):
    """
    Calcola media/min/max/count dei pixel dentro un poligono.
    Il poligono è in [lat, lon]; va convertito a [lon, lat] per shapely.
    Ritorna dict o None se non ci sono pixel validi.
    """
    # Shapely usa (x=lon, y=lat)
    poly_shapely = Polygon([(lon, lat) for lat, lon in polygon_latlon])
    geojson = mapping(poly_shapely)

    with rasterio.open(io.BytesIO(tiff_bytes)) as src:
        try:
            masked, _ = rasterio.mask.mask(src, [geojson], crop=True, nodata=src.nodata)
        except ValueError as e:
            # Poligono fuori dal raster
            log.warning(f'  mask fallito: {e}')
            return None

    arr = masked[0]  # primo banda
    nodata = src.nodata if src.nodata is not None else -9999

    # Costruisce maschera dei pixel validi
    valid_mask = (arr != nodata) & np.isfinite(arr) & (arr > -900)
    valid_values = arr[valid_mask]

    if valid_values.size == 0:
        return {'mean': None, 'min': None, 'max': None, 'pixel_count': 0}

    return {
        'mean':        float(np.mean(valid_values)),
        'min':         float(np.min(valid_values)),
        'max':         float(np.max(valid_values)),
        'pixel_count': int(valid_values.size),
    }


def value_at_point(tiff_bytes, lat, lon):
    """Restituisce il valore del pixel alla coordinata (lat, lon), o None."""
    with rasterio.open(io.BytesIO(tiff_bytes)) as src:
        try:
            row, col = src.index(lon, lat)
            if row < 0 or col < 0 or row >= src.height or col >= src.width:
                return None
            val = src.read(1)[row, col]
        except Exception:
            return None

    nodata = src.nodata if src.nodata is not None else -9999
    if val == nodata or not np.isfinite(val) or val < -900:
        return None
    return float(val)

# ─── Generazione timestamps da archiviare ─────────────────────────────────────

def timestamps_for_day(product, day_utc):
    """
    Ritorna lista di datetime UTC per il prodotto in un dato giorno.
    Per CUM24: [day 00:00 UTC]
    Per CUM3:  [day 03, 06, 09, 12, 15, 18, 21, day+1 00] UTC
    """
    out = []
    for h in PRODUCTS[product]['utc_hours_in_day']:
        if h == 24:
            ts = datetime(day_utc.year, day_utc.month, day_utc.day,
                          tzinfo=timezone.utc) + timedelta(days=1)
        else:
            ts = datetime(day_utc.year, day_utc.month, day_utc.day, h,
                          tzinfo=timezone.utc)
        out.append(ts)
    return out

# ─── CSV idempotente ─────────────────────────────────────────────────────────

def _csv_path(data_dir, area_name, product):
    return data_dir / f'{area_name}_{product.lower()}.csv'


def load_existing_keys(csv_path):
    """Carica le chiavi (timestamp_utc, location_name) già presenti nel CSV."""
    if not csv_path.exists():
        return set()
    keys = set()
    with open(csv_path, newline='', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            keys.add((row['timestamp_utc'], row['location_name']))
    return keys


def append_rows(csv_path, rows):
    """Scrive le righe in append. Crea l'header se il file non esiste."""
    if not rows:
        return
    write_header = not csv_path.exists()
    with open(csv_path, 'a', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=CSV_HEADERS)
        if write_header:
            writer.writeheader()
        writer.writerows(rows)


def replace_keys(csv_path, keys_to_remove):
    """Rimuove dal CSV le righe con le date chiavi specificate (per --force)."""
    if not csv_path.exists():
        return
    kept = []
    with open(csv_path, newline='', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            key = (row['timestamp_utc'], row['location_name'])
            if key not in keys_to_remove:
                kept.append(row)
    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=CSV_HEADERS)
        writer.writeheader()
        writer.writerows(kept)

# ─── Generazione XLSX da CSV ─────────────────────────────────────────────────

def regenerate_xlsx(data_dir, area_name):
    """Crea/sovrascrive {area}.xlsx con due sheet: CUM24 e CUM3."""
    wb = Workbook()
    wb.remove(wb.active)  # rimuove sheet di default

    for product in ['CUM24', 'CUM3']:
        csv_p = _csv_path(data_dir, area_name, product)
        ws = wb.create_sheet(product)
        if not csv_p.exists():
            ws.append(CSV_HEADERS)
            continue
        with open(csv_p, newline='', encoding='utf-8') as f:
            reader = csv.reader(f)
            for row in reader:
                ws.append(row)
        # Auto-width approssimato (prime 6 colonne)
        for col_idx in range(1, 7):
            ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = 18

    xlsx_path = data_dir / f'{area_name}.xlsx'
    wb.save(xlsx_path)
    log.info(f'  → {xlsx_path.name}')

# ─── Pipeline principale ─────────────────────────────────────────────────────

def process_day_product(area, product, day, data_dir, force=False):
    """
    Processa un giorno per un'area+prodotto.
    Aggiunge righe al CSV idempotentemente.
    Ritorna numero di righe nuove scritte.
    """
    csv_path = _csv_path(data_dir, area['name'], product)
    existing = load_existing_keys(csv_path)

    timestamps = timestamps_for_day(product, day)
    rows_to_add = []
    keys_to_force_remove = set()
    now_iso = datetime.now(tz=timezone.utc).isoformat(timespec='seconds').replace('+00:00', 'Z')

    for ts in timestamps:
        ts_iso = ts.isoformat().replace('+00:00', 'Z')
        ts_ms = int(ts.timestamp() * 1000)

        # Chiavi che useremo
        area_key = (ts_iso, area['name'])
        vertex_keys = [(ts_iso, f"{area['name']}_{v['id']}") for v in area['sample_vertices']]

        if not force:
            # Skippa se già presenti TUTTE le chiavi attese
            all_present = area_key in existing and all(k in existing for k in vertex_keys)
            if all_present:
                continue
        else:
            keys_to_force_remove.add(area_key)
            keys_to_force_remove.update(vertex_keys)

        log.info(f'  → fetch {product} @ {ts_iso}')
        url = get_pre_signed_url(product, ts_ms)
        if not url:
            continue
        tiff = download_geotiff(url)
        if not tiff:
            continue

        # Stats area
        area_stats = stats_for_polygon(tiff, area['polygon'])
        if area_stats is None:
            log.warning(f'  area stats N/D per {area["name"]} @ {ts_iso}')
            continue

        rows_to_add.append({
            'timestamp_utc':   ts_iso,
            'product':         product,
            'area_name':       area['name'],
            'location_type':   'area',
            'location_name':   area['name'],
            'lat':             '',
            'lon':             '',
            'value':           '',
            'mean':            f"{area_stats['mean']:.3f}" if area_stats['mean'] is not None else '',
            'min':             f"{area_stats['min']:.3f}"  if area_stats['min']  is not None else '',
            'max':             f"{area_stats['max']:.3f}"  if area_stats['max']  is not None else '',
            'pixel_count':     area_stats['pixel_count'],
            'fetched_at_utc':  now_iso,
        })

        # Estrazione puntuale sui vertici
        for v in area['sample_vertices']:
            val = value_at_point(tiff, v['lat'], v['lon'])
            rows_to_add.append({
                'timestamp_utc':   ts_iso,
                'product':         product,
                'area_name':       area['name'],
                'location_type':   'vertex',
                'location_name':   f"{area['name']}_{v['id']}",
                'lat':             f"{v['lat']:.6f}",
                'lon':             f"{v['lon']:.6f}",
                'value':           f"{val:.3f}" if val is not None else '',
                'mean':            '',
                'min':             '',
                'max':             '',
                'pixel_count':     '',
                'fetched_at_utc':  now_iso,
            })

    if force and keys_to_force_remove:
        replace_keys(csv_path, keys_to_force_remove)

    append_rows(csv_path, rows_to_add)
    return len(rows_to_add)


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument('--days', type=int, default=1,
                        help='Numero di giorni da recuperare a partire da ieri (default 1, max 30).')
    parser.add_argument('--force', action='store_true',
                        help='Rimpiazza le righe esistenti invece di skipparle.')
    parser.add_argument('--areas-file', default=None,
                        help='Path alternativo al file areas.json.')
    parser.add_argument('--data-dir', default=None,
                        help='Cartella di output (default: archive/data accanto allo script).')
    args = parser.parse_args()

    days = max(1, min(args.days, 30))

    script_dir = Path(__file__).resolve().parent
    archive_dir = script_dir.parent
    areas_file = Path(args.areas_file) if args.areas_file else archive_dir / 'areas.json'
    data_dir   = Path(args.data_dir)   if args.data_dir   else archive_dir / 'data'
    data_dir.mkdir(parents=True, exist_ok=True)

    log.info(f'Loading areas from {areas_file}')
    config = json.loads(areas_file.read_text(encoding='utf-8'))
    areas = config['areas']
    log.info(f'Configured areas: {[a["label"] for a in areas]}')

    # Lavoreremo "ieri" a ritroso di N-1 giorni: oggi non è ancora completo
    today_utc = datetime.now(tz=timezone.utc).date()
    days_to_process = [today_utc - timedelta(days=i + 1) for i in range(days)]
    days_to_process.reverse()  # ordine cronologico

    log.info(f'Processing days: {[d.isoformat() for d in days_to_process]}')

    total_new = 0
    for day in days_to_process:
        log.info(f'─── Day {day.isoformat()} ───────────────────────')
        for area in areas:
            for product in PRODUCTS.keys():
                log.info(f'[{area["label"]}] {product}')
                n = process_day_product(area, product, day, data_dir, force=args.force)
                log.info(f'  + {n} new rows')
                total_new += n

    # Rigenera tutti gli XLSX
    log.info('─── Rebuilding XLSX ────────────────────────')
    for area in areas:
        regenerate_xlsx(data_dir, area['name'])

    log.info(f'Done. Total new rows: {total_new}')
    return 0


if __name__ == '__main__':
    try:
        sys.exit(main())
    except KeyboardInterrupt:
        log.info('Interrotto.')
        sys.exit(130)
    except Exception as e:
        log.error(f'FATAL: {e}', exc_info=True)
        sys.exit(1)
